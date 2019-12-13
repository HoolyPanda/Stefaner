import cryptor as Cryptor
import openpyxl


# DEMO
a = Cryptor.Cryptor()
a.getDictionary()

book = openpyxl.load_workbook(filename = "./test.xlsx")
thoughts = book.get_sheet_by_name(book.get_sheet_names()[0])
output = book.get_sheet_by_name(book.get_sheet_names()[1])
rows = thoughts.max_row
columns = thoughts.max_column
texts = []

for row in range(1, rows + 1):
    for column in range(1, columns + 1):
        cellValue = thoughts.cell(row, column).value
        if cellValue:
            texts.append(cellValue)
    if texts:
        a.Encrypt(text1 = texts[0], text2 = texts[1], text3 = texts[2])
        a.Decrypt(encryptedText= a.et, key= a.keys[0])
        a.Decrypt(encryptedText= a.et, key= a.keys[1])
        a.Decrypt(encryptedText= a.et, key= a.keys[2])
        for column in range(1, columns + 1):
            t = a.keys[column - 1]
            output.cell(row, column).value = t
        output.cell(row, 4).value = a.et
        book.save(filename = "./test.xlsx")
        texts = []

